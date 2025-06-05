from flask import Flask, render_template, request
import pandas as pd
import os
from openpyxl import load_workbook
import errno

app = Flask(__name__)


def get_total_participants():
    fichier_excel = "reponse.xlsx"
    try:
        if os.path.exists(fichier_excel):
            df = pd.read_excel(fichier_excel)
            return df[df['Présence'] == 'Oui']['Nombre de personnes'].sum()
        return 0
    except Exception as e:
        print(f"Erreur lors de la lecture du total: {e}")
        return 0


def sauvegarder_excel(nom, prenom, telephone, presence, nombre_personnes):
    fichier_excel = "reponse.xlsx"

    try:
        # S'assurer que nombre_personnes est un entier
        nombre_personnes = int(nombre_personnes)

        # Créer un nouveau DataFrame avec la réponse
        nouvelle_reponse = pd.DataFrame({
            'Nom': [nom],
            'Prénom': [prenom],
            'Téléphone': [telephone],
            'Présence': [presence],
            'Nombre de personnes': [nombre_personnes if presence == 'Oui' else 0]
        })

        print(f"Debug - Enregistrement: Nom={nom}, Prénom={prenom}, Téléphone={telephone}, Présence={presence}, Nombre={nombre_personnes}")

        # Si le fichier existe, lire et ajouter la nouvelle réponse
        if os.path.exists(fichier_excel):
            try:
                df_existant = pd.read_excel(fichier_excel)
            except PermissionError:
                return False, "Le fichier Excel est actuellement ouvert. Veuillez le fermer et réessayer."
            df_final = pd.concat([df_existant, nouvelle_reponse], ignore_index=True)
        else:
            df_final = nouvelle_reponse

        try:
            # Sauvegarder dans Excel
            df_final.to_excel(fichier_excel, index=False)

            # Calculer le total des personnes présentes
            total_presents = df_final[df_final['Présence'] == 'Oui']['Nombre de personnes'].sum()
            print(f"Debug - Total après sauvegarde: {total_presents}")

            # Utiliser openpyxl pour modifier la cellule D2
            wb = load_workbook(fichier_excel)
            ws = wb.active
            ws['D1'] = 'Total Participants'
            ws['D2'] = total_presents
            wb.save(fichier_excel)

            return True, total_presents
        except PermissionError:
            return False, "Le fichier Excel est actuellement ouvert. Veuillez le fermer et réessayer."

    except Exception as e:
        print(f"Erreur lors de la sauvegarde: {e}")
        return False, "Une erreur est survenue lors de l'enregistrement."


@app.route('/')
def index():
    total_participants = get_total_participants()
    return render_template('index.html', total_participants=total_participants)


@app.route('/submit', methods=['POST'])
def submit():
    nom = request.form.get('nom')
    prenom = request.form.get('prenom')
    telephone = request.form.get('telephone')
    presence = request.form.get('presence')
    nombre_personnes = request.form.get('nombre_personnes', '0')

    print(f"Debug - Données reçues: {request.form}")

    if not all([nom, prenom, telephone, presence]):
        return render_template('index.html',
                               message="Veuillez remplir tous les champs",
                               message_type="error",
                               total_participants=get_total_participants())

    # Validation du numéro de téléphone (10 chiffres)
    if not telephone.isdigit() or len(telephone) != 10:
        return render_template('index.html',
                               message="Le numéro de téléphone doit contenir 10 chiffres",
                               message_type="error",
                               total_participants=get_total_participants())

    try:
        nombre_personnes = int(nombre_personnes)
        if presence == 'Oui' and nombre_personnes < 1:
            raise ValueError("Le nombre de personnes doit être au moins 1 pour une réponse positive")
    except ValueError as e:
        print(f"Debug - Erreur de validation: {e}")
        return render_template('index.html',
                               message="Le nombre de personnes doit être un nombre valide",
                               message_type="error",
                               total_participants=get_total_participants())

    succes, message = sauvegarder_excel(nom, prenom, telephone, presence, nombre_personnes)
    if succes:
        return render_template('index.html',
                               message="Votre réponse a été enregistrée.",
                               message_type="success",
                               total_participants=message)
    else:
        return render_template('index.html',
                               message=message,
                               message_type="error",
                               total_participants=get_total_participants())


if __name__ == '__main__':
    app.run(debug=True)